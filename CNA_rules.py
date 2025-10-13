# CNA_Rules.py
import re
import pandas as pd
from CNA_utils import (
    base_dir, ensure_datetime, compute_dly_real, compute_dly_wo_handling,
    write_excel, highlight_rows_by_nonempty, highlight_rows_by_threshold, HANDLING_CODES
)


def etihad(df: pd.DataFrame, filename: str = "FCO_Delays_ETHIAD.xlsx") -> str:
    """
    Partenze EY/ETIHAD/ETHIAD, calcolo/uso DLY_REAL (min, >0), ordinamento per STD asc,
    Excel con righe evidenziate se DLY_REAL > 60.
    """
    out = df.copy()
    out["A/D"] = out["A/D"].astype(str).str.strip().str.upper()
    out["IATA"] = out["IATA"].astype(str).str.strip().str.upper()

    out = out[out["A/D"].eq("D") & out["IATA"].isin({"EY", "ETIHAD", "ETHIAD"})]
    if out.empty:
        print("Nessuna partenza per EY/ETIHAD/ETHIAD. Nessun file creato.")
        return ""

    out = ensure_datetime(out, ["STD", "ATD"])
    if "DLY_REAL" not in out.columns or not pd.api.types.is_numeric_dtype(out["DLY_REAL"]):
        out = compute_dly_real(out, "ATD", "STD", "DLY_REAL")

    out = out.sort_values("STD", ascending=True, na_position="last").reset_index(drop=True)

    def _hl(ws, df_):
        highlight_rows_by_threshold(ws, df_, "DLY_REAL", 60, color="FFFFFF00")

    path = write_excel(out, filename, sheet="EY_D", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path


def united(df: pd.DataFrame, filename: str = "FCO_Delays_UNITED.xlsx") -> str:
    """
    Allinea A (arrivo) e D (partenza) per IATA='UA' su ID; calcola DLY_REAL (min, >0),
    ADV_IN (min anticipo arrivo), %TURN_RATE_IN/OUT, DLY_WO_HNDLG e INFO_REQUIRED.
    Ordina per STD ascendente. Evidenzia percentuali > 0% e celle DLY_1/DLY_2 con codici handling.
    """
    req = ["ID","A/D","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
           "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Colonne mancanti nel DataFrame di input: {miss}")

    base = df.copy()
    base["IATA"] = base["IATA"].astype(str).str.strip().str.upper()
    base["A/D"]  = base["A/D"].astype(str).str.strip().str.upper()
    base = base[base["IATA"].eq("UA") & base["A/D"].isin(["A","D"])]
    if base.empty:
        print("Nessuna riga con IATA='UA' e A/D in {A,D}. Nessun file creato.")
        return ""

    base = ensure_datetime(base, ["STD","ATD"])

    # Arrivi
    A = (
        base[base["A/D"].eq("A")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","IATA","FLT_N","STD","ATD"]]
        .rename(columns={
            "TRANSPORT":"TRANSPORT_A",
            "FLT_TYPE":"FLT_TYPE_A",
            "IATA":"IATA_IN",
            "FLT_N":"FLT_IN",
            "STD":"STA",
            "ATD":"ATA",
        })
    )

    # Partenze
    D = (
        base[base["A/D"].eq("D")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","FROM","TO","IATA","FLT_N",
                 "STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]]
        .rename(columns={
            "TRANSPORT":"TRANSPORT_D",
            "FLT_TYPE":"FLT_TYPE_D",
            "IATA":"IATA_OUT",
            "FLT_N":"FLT_OUT",
        })
    )

    if A.empty or D.empty:
        print("Mancano arrivi o partenze UA per effettuare l'allineamento. Nessun file creato.")
        return ""

    out = pd.merge(A, D, on="ID", how="inner")
    if out.empty:
        print("Nessuna coppia A/D con stesso ID per UA. Nessun file creato.")
        return ""

    # DLY_REAL sulle PARTENZE
    out = compute_dly_real(out, "ATD", "STD", "DLY_REAL")

    # ADV_IN (min anticipo arrivo)
    adv_min = (out["STA"] - out["ATA"]).dt.total_seconds().div(60)
    out["ADV_IN"] = adv_min.mask(adv_min < 0, 0).round().astype("Int64")
    out.insert(out.columns.get_loc("ATA") + 1, "ADV_IN", out.pop("ADV_IN"))

    # %TURN_RATE_IN da ADV_IN
    def pct_from_minutes(m):
        if pd.isna(m): return "0%"
        m = int(m)
        if m <= 60:   return "0%"
        if m <= 120:  return "15%"
        if m <= 180:  return "25%"
        if m <= 240:  return "50%"
        return "100%"

    out["%TURN_RATE_IN"]  = out["ADV_IN"].map(pct_from_minutes)

    # DLY_WO_HNDLG e %_TURN_RATE_OUT
    out = compute_dly_wo_handling(out, "DLY_REAL", "DLY_1","DLY_1_t","DLY_2","DLY_2_t",
                                  handling_codes=HANDLING_CODES, out_col="DLY_WO_HNDLG")
    out["%_TURN_RATE_OUT"] = out["DLY_WO_HNDLG"].map(pct_from_minutes)

    # INFO_REQUIRED: "YES" se DLY_1_t + DLY_2_t != DLY_REAL (dove DLY_REAL è disponibile)
    d1_min = pd.to_numeric(out["DLY_1_t"], errors="coerce").fillna(0)
    d2_min = pd.to_numeric(out["DLY_2_t"], errors="coerce").fillna(0)
    imput_sum = (d1_min + d2_min)
    dly_real_num = pd.to_numeric(out["DLY_REAL"], errors="coerce")
    out["INFO_REQUIRED"] = pd.Series(
        pd.NA if pd.isna(rv) else ("YES" if int(round(imput_sum.iloc[i])) != int(round(rv)) else "")
        for i, rv in enumerate(dly_real_num)
    )

    # Ordine finale e ordinamento
    final_cols = [
        "ID","TRANSPORT_A","TRANSPORT_D","FLT_TYPE_A","FLT_TYPE_D","REG","MOD","MTOW","STAND",
        "FROM","TO","IATA_IN","FLT_IN","STA","ATA","ADV_IN","IATA_OUT","FLT_OUT","STD","ATD",
        "DLY_REAL","DLY_WO_HNDLG","DLY_1","DLY_1_t","DLY_2","DLY_2_t",
        "%TURN_RATE_IN","%_TURN_RATE_OUT","INFO_REQUIRED"
    ]
    out = out.loc[:, final_cols]
    out = out.sort_values("STD", ascending=True, na_position="last").reset_index(drop=True)

    def _hl(ws, df_):
        # evidenzia % > 0% e celle DLY_1/DLY_2 con codici handling
        from openpyxl.styles import PatternFill
        yellow = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
        cols = list(df_.columns)
        in_col   = cols.index("%TURN_RATE_IN") + 1
        out_col  = cols.index("%_TURN_RATE_OUT") + 1
        dly1_col = cols.index("DLY_1") + 1
        dly2_col = cols.index("DLY_2") + 1
        max_row, max_col = ws.max_row, ws.max_column
        for r in range(2, max_row + 1):
            v_in  = ws.cell(row=r, column=in_col).value
            v_out = ws.cell(row=r, column=out_col).value
            if isinstance(v_in, str) and v_in.strip() != "0%":
                ws.cell(row=r, column=in_col).fill = yellow
            if isinstance(v_out, str) and v_out.strip() != "0%":
                ws.cell(row=r, column=out_col).fill = yellow
            # codici handling nelle celle DLY_1/DLY_2
            for c in (dly1_col, dly2_col):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                m = re.search(r"\d+", str(v))
                if m and (int(m.group(0)) in HANDLING_CODES):
                    ws.cell(row=r, column=c).fill = yellow

    path = write_excel(out, filename, sheet="TURN_RATES_UA", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path


def delta(df: pd.DataFrame, filename: str = "FCO_Delays_DELTA.xlsx") -> str:
    """
    Allinea A e D per IATA='DL' su ID; calcola DLY_REAL; definisce SURCHARGE:
      - 30% se DLY_REAL>180 e 07:00≤ATD≤21:00 e nessuno dei due FLT_TYPE è 'FERRY'
      - 15% se DLY_REAL>180 e 07:00≤ATD≤21:00 e almeno uno è 'FERRY'
    Ordina per STD asc, evidenzia le righe con SURCHARGE valorizzato.
    """
    req = ["ID","A/D","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
           "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Colonne mancanti: {miss}")

    base = df.copy()
    base["IATA"] = base["IATA"].astype(str).str.strip().str.upper()
    base["A/D"]  = base["A/D"].astype(str).str.strip().str.upper()
    base = base[base["IATA"].eq("DL") & base["A/D"].isin(["A","D"])]
    if base.empty:
        print("Nessuna riga DL con A/D in {A,D}. Nessun file creato.")
        return ""

    base = ensure_datetime(base, ["STD","ATD"])

    # Arrivi
    A = (
        base[base["A/D"].eq("A")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","FLT_N","STD","ATD"]]
        .rename(columns={
            "TRANSPORT":"TRANSPORT_A",
            "FLT_TYPE":"FLT_TYPE_A",
            "FLT_N":"FLT_N_IN",
            "STD":"STA",
            "ATD":"ATA",
        })
    )

    # Partenze
    D = (
        base[base["A/D"].eq("D")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
                 "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]]
        .rename(columns={
            "TRANSPORT":"TRANSPORT_D",
            "FLT_TYPE":"FLT_TYPE_D",
            "FLT_N":"FLT_N_OUT",
        })
    )

    if A.empty or D.empty:
        print("Mancano arrivi o partenze DL per effettuare l'allineamento. Nessun file creato.")
        return ""

    D = compute_dly_real(D, "ATD", "STD", "DLY_REAL")
    out = pd.merge(A, D, on="ID", how="inner")
    if out.empty:
        print("Nessuna coppia A/D con stesso ID per DL. Nessun file creato.")
        return ""

    out = out.sort_values("STD", ascending=True, na_position="last").reset_index(drop=True)

    # Surcharge rules (usano ATD come orario finestra, coerente con tua versione)
    std_hour = out["ATD"].dt.hour
    in_window = out["ATD"].notna() & std_hour.between(7, 21, inclusive="both")
    dly_ok = out["DLY_REAL"].notna() & (out["DLY_REAL"] > 180)
    ferry_ok = out["FLT_TYPE_A"].astype(str).str.upper().ne("FERRY") & \
               out["FLT_TYPE_D"].astype(str).str.upper().ne("FERRY")

    out["SURCHARGE"] = ""
    base_cond = dly_ok & in_window
    out.loc[base_cond & ferry_ok,    "SURCHARGE"] = "30%"
    out.loc[base_cond & (~ferry_ok), "SURCHARGE"] = "15%"

    cols = [
        "ID","TRANSPORT_A","TRANSPORT_D","FLT_TYPE_A","FLT_TYPE_D","REG","MOD","MTOW","STAND",
        "IATA","FROM","TO","FLT_N_IN","STA","ATA","FLT_N_OUT","STD","ATD",
        "DLY_REAL","DLY_1","DLY_1_t","DLY_2","DLY_2_t","SURCHARGE"
    ]
    out = out.loc[:, cols]

    def _hl(ws, df_):
        highlight_rows_by_nonempty(ws, df_, "SURCHARGE", color="FFFFFF00")

    path = write_excel(out, filename, sheet="DL_AD", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path


def ritardo_generico(df: pd.DataFrame, iata_code: str, min_minutes: int,
                     filename: str | None = None) -> str:
    """
    Partenze per IATA specifico, join con eventuale arrivo, DLY_REAL e DLY_WO_HNDLG,
    ordinamento per STD asc; evidenzia righe con DLY_WO_HNDLG ≥ min_minutes.
    """
    req = ["ID","A/D","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
           "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Colonne mancanti nel DataFrame di input: {miss}")

    iata = str(iata_code).strip().upper()
    if filename is None:
        filename = f"FCO_Delays_{iata}_{int(min_minutes)}.xlsx"

    base = df.copy()
    base["IATA"] = base["IATA"].astype(str).str.strip().str.upper()
    base["A/D"]  = base["A/D"].astype(str).str.strip().str.upper()
    base = ensure_datetime(base, ["STD","ATD"])

    D = (
        base[base["A/D"].eq("D") & base["IATA"].eq(iata)]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
                 "FROM","TO","FLT_N","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_D","FLT_TYPE":"FLT_TYPE_D","FLT_N":"FLT_N_OUT"})
    )
    if D.empty:
        print(f"Nessuna partenza per IATA='{iata}'. Nessun file creato.")
        return ""

    A = (
        base[base["A/D"].eq("A")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","FLT_N","STD","ATD"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_A","FLT_TYPE":"FLT_TYPE_A",
                         "FLT_N":"FLT_N_IN","STD":"STA","ATD":"ATA"})
    )

    out = pd.merge(D, A, on="ID", how="left")
    out = compute_dly_real(out, "ATD", "STD", "DLY_REAL")
    out = compute_dly_wo_handling(out, "DLY_REAL", "DLY_1","DLY_1_t","DLY_2","DLY_2_t",
                                  handling_codes=HANDLING_CODES, out_col="DLY_WO_HNDLG")
    out = out.sort_values("STD", ascending=True, na_position="last").reset_index(drop=True)

    final_cols = [
        "ID","TRANSPORT_A","TRANSPORT_D","FLT_TYPE_A","FLT_TYPE_D","REG","MOD","MTOW","STAND",
        "IATA","FROM","TO","FLT_N_IN","STA","ATA","FLT_N_OUT","STD","ATD",
        "DLY_REAL","DLY_1","DLY_1_t","DLY_2","DLY_2_t","DLY_WO_HNDLG"
    ]
    for c in final_cols:
        if c not in out.columns: out[c] = pd.NA
    out = out.loc[:, final_cols]

    def _hl(ws, df_):
        highlight_rows_by_threshold(ws, df_, "DLY_WO_HNDLG", min_minutes, color="FFFFFF00")

    path = write_excel(out, filename, sheet=f"{iata}_D_{int(min_minutes)}", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path


def arkia(df: pd.DataFrame, filename: str = "FCO_Delays_ARKIA.xlsx") -> str:
    """
    Partenze IATA='IZ', join con arrivo, DLY_REAL, DLY_WO_HNDLG, SURCHARGE per scaglioni
    (20% 91–120, 30% 121–180, 45% >180); evidenzia righe con SURCHARGE.
    """
    req = ["ID","A/D","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
           "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Colonne mancanti nel DataFrame di input: {miss}")

    base = df.copy()
    base["IATA"] = base["IATA"].astype(str).str.strip().str.upper()
    base["A/D"]  = base["A/D"].astype(str).str.strip().str.upper()
    base = ensure_datetime(base, ["STD","ATD"])

    D = (
        base[base["A/D"].eq("D") & base["IATA"].eq("IZ")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
                 "FROM","TO","FLT_N","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_D","FLT_TYPE":"FLT_TYPE_D","FLT_N":"FLT_N_OUT"})
    )
    if D.empty:
        print("Nessuna partenza per IATA='IZ'. Nessun file creato.")
        return ""

    A = (
        base[base["A/D"].eq("A")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","FLT_N","STD","ATD"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_A","FLT_TYPE":"FLT_TYPE_A",
                         "FLT_N":"FLT_N_IN","STD":"STA","ATD":"ATA"})
    )

    out = pd.merge(D, A, on="ID", how="left")
    out = compute_dly_real(out, "ATD", "STD", "DLY_REAL")
    out = compute_dly_wo_handling(out, "DLY_REAL", "DLY_1","DLY_1_t","DLY_2","DLY_2_t",
                                  handling_codes=HANDLING_CODES, out_col="DLY_WO_HNDLG")

    def surcharge_from_minutes(m):
        if pd.isna(m): return ""
        m = int(m)
        if 91 <= m <= 120: return "20%"
        if 121 <= m <= 180: return "30%"
        if m > 180:        return "45%"
        return ""

    out["SURCHARGE"] = out["DLY_WO_HNDLG"].map(surcharge_from_minutes)
    out = out.sort_values("STD", ascending=True, na_position="last").reset_index(drop=True)

    final_cols = [
        "ID","TRANSPORT_A","TRANSPORT_D","FLT_TYPE_A","FLT_TYPE_D","REG","MOD","MTOW","STAND",
        "IATA","FROM","TO","FLT_N_IN","STA","ATA","FLT_N_OUT","STD","ATD",
        "DLY_REAL","DLY_1","DLY_1_t","DLY_2","DLY_2_t","DLY_WO_HNDLG","SURCHARGE"
    ]
    for c in final_cols:
        if c not in out.columns: out[c] = pd.NA
    out = out.loc[:, final_cols]

    def _hl(ws, df_):
        highlight_rows_by_nonempty(ws, df_, "SURCHARGE", color="FFFFFF00")

    path = write_excel(out, filename, sheet="IZ_D", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path


def anticipo_generico(df: pd.DataFrame, iata_code: str, min_minutes: int,
                      filename: str | None = None) -> str:
    """
    Arrivi per IATA specifico, join con eventuale partenza, calcolo ADV_IN (min anticipo),
    ordinamento per STA asc, evidenzia righe con ADV_IN ≥ min_minutes.
    """
    req = ["ID","A/D","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","IATA",
           "FLT_N","FROM","TO","STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Colonne mancanti nel DataFrame di input: {miss}")

    iata = str(iata_code).strip().upper()
    if filename is None:
        filename = f"FCO_Early_{iata}_{int(min_minutes)}.xlsx"

    base = df.copy()
    base["IATA"] = base["IATA"].astype(str).str.strip().str.upper()
    base["A/D"]  = base["A/D"].astype(str).str.strip().str.upper()
    base = ensure_datetime(base, ["STD","ATD"])

    A = (
        base[base["A/D"].eq("A") & base["IATA"].eq(iata)]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","IATA","FLT_N","STD","ATD"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_A","FLT_TYPE":"FLT_TYPE_A",
                         "FLT_N":"FLT_N_IN","STD":"STA","ATD":"ATA"})
    )
    if A.empty:
        print(f"Nessun arrivo per IATA='{iata}'. Nessun file creato.")
        return ""

    D = (
        base[base["A/D"].eq("D")]
        .sort_values(["ID","STD"]).drop_duplicates(subset=["ID"], keep="last")
        .loc[:, ["ID","TRANSPORT","FLT_TYPE","REG","MOD","MTOW","STAND","FROM","TO","FLT_N",
                 "STD","ATD","DLY_1","DLY_1_t","DLY_2","DLY_2_t"]]
        .rename(columns={"TRANSPORT":"TRANSPORT_D","FLT_TYPE":"FLT_TYPE_D","FLT_N":"FLT_N_OUT"})
    )

    out = pd.merge(A, D, on="ID", how="left")

    adv_min = (out["STA"] - out["ATA"]).dt.total_seconds().div(60)
    out["ADV_IN"] = adv_min.mask(adv_min < 0, 0).round().astype("Int64")

    out = out.sort_values("STA", ascending=True, na_position="last").reset_index(drop=True)

    final_cols = [
        "ID","TRANSPORT_A","TRANSPORT_D","FLT_TYPE_A","FLT_TYPE_D","REG","MOD","MTOW","STAND",
        "IATA","FROM","TO","FLT_N_IN","STA","ATA","ADV_IN","FLT_N_OUT","STD","ATD",
        "DLY_1","DLY_1_t","DLY_2","DLY_2_t"
    ]
    for c in final_cols:
        if c not in out.columns: out[c] = pd.NA
    out = out.loc[:, final_cols]

    def _hl(ws, df_):
        highlight_rows_by_threshold(ws, df_, "ADV_IN", min_minutes, color="FFFFFF00")

    path = write_excel(out, filename, sheet=f"{iata}_A_{int(min_minutes)}", highlighter=_hl)
    print(f"File Excel creato: {path}  (righe: {out.shape[0]})")
    return path

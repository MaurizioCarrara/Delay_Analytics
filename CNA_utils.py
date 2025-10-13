# CNA_utils.py
import os
import sys
import pandas as pd
from openpyxl.styles import PatternFill

# Codici ritardo attribuibili all'handling (da sottrarre nel "senza handling")
HANDLING_CODES = {12, 13, 15, 18, 31, 32, 33, 34, 35, 39, 52}


def base_dir() -> str:
    """
    Restituisce la cartella dell'eseguibile (se PyInstaller onefile),
    altrimenti la cartella del file .py.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "executable"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def ensure_datetime(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """Converte le colonne indicate in datetime (dayfirst=True) se non lo sono già."""
    for c in cols:
        if c in df.columns and not pd.api.types.is_datetime64_any_dtype(df[c]):
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
    return df


def compute_dly_real(df: pd.DataFrame,
                     atd_col: str = "ATD",
                     std_col: str = "STD",
                     out_col: str = "DLY_REAL") -> pd.DataFrame:
    """Calcola il ritardo reale in minuti (ATD-STD), solo positivi, in out_col (Int64)."""
    mins = (df[atd_col] - df[std_col]).dt.total_seconds().div(60)
    df[out_col] = mins.where(mins > 0).round().astype("Int64")
    return df


def compute_dly_wo_handling(df: pd.DataFrame,
                            dly_real_col: str = "DLY_REAL",
                            dly1_col: str = "DLY_1", dly1_min_col: str = "DLY_1_t",
                            dly2_col: str = "DLY_2", dly2_min_col: str = "DLY_2_t",
                            handling_codes: set[int] = HANDLING_CODES,
                            out_col: str = "DLY_WO_HNDLG") -> pd.DataFrame:
    """
    Calcola il ritardo 'senza handling': DLY_REAL meno i minuti associati ai codici handling
    presenti in DLY_1/DLY_2. Risultato (Int64) clip ≥ 0.
    """
    d1_code = pd.to_numeric(df.get(dly1_col), errors="coerce").astype("Int64")
    d2_code = pd.to_numeric(df.get(dly2_col), errors="coerce").astype("Int64")
    d1_min  = pd.to_numeric(df.get(dly1_min_col), errors="coerce").fillna(0)
    d2_min  = pd.to_numeric(df.get(dly2_min_col), errors="coerce").fillna(0)

    sub_d1 = d1_min.where(d1_code.isin(handling_codes), 0)
    sub_d2 = d2_min.where(d2_code.isin(handling_codes), 0)

    dly_real_num = pd.to_numeric(df.get(dly_real_col), errors="coerce").fillna(0)
    dly_wo = (dly_real_num - sub_d1 - sub_d2).clip(lower=0)
    df[out_col] = dly_wo.round().astype("Int64")
    return df


# ============================== Excel writer + evidenziazioni ==============================

def write_excel(df: pd.DataFrame, filename: str, sheet: str,
                datetime_fmt: str = "DD-MM-YYYY hh:mm",
                date_fmt: str = "DD-MM-YYYY",
                highlighter=None) -> str:
    """
    Scrive df in Excel (cartella accanto all’eseguibile o al .py) e, se passato,
    applica una funzione di evidenziazione:
      highlighter: callable(ws, df) -> None
    """
    out_path = os.path.join(base_dir(), filename)
    with pd.ExcelWriter(out_path, engine="openpyxl",
                        datetime_format=datetime_fmt, date_format=date_fmt) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        if highlighter is not None:
            ws = writer.sheets[sheet]
            highlighter(ws, df)
    return out_path


def highlight_rows_by_nonempty(ws, df: pd.DataFrame, col_name: str, color: str = "FFFFFF00"):
    """Evidenzia INTERA RIGA se la cella col_name non è vuota."""
    fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
    col_idx = df.columns.get_loc(col_name) + 1  # 1-based
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip() != "":
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = fill


def highlight_rows_by_threshold(ws, df: pd.DataFrame, col_name: str, threshold: float,
                                color: str = "FFFFFF00"):
    """Evidenzia INTERA RIGA se il valore numerico in col_name è ≥ threshold."""
    fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
    col_idx = df.columns.get_loc(col_name) + 1  # 1-based
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        try:
            val = float(v)
        except (TypeError, ValueError):
            try:
                val = float(str(v).replace(",", "."))
            except Exception:
                val = float("-inf")
        if val >= float(threshold):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = fill
